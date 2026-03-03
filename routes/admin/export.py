# routes/admin/export.py
from flask import Blueprint, send_file
from openpyxl import Workbook
from io import BytesIO
from flask_babel import gettext as _
from models import Customer, Product, CashDebt, InstallmentProduct, Sale
from utils.admin_auth import admin_required

export_bp = Blueprint("export", __name__, url_prefix="/admin/export")


@export_bp.get("/excel")
@admin_required
def export_excel():
    wb = Workbook()

    ws = wb.active
    ws.title = "Customers"
    ws.append(["ID", "Custom ID", "Name", "Phone", "Ledger"])

    for c in Customer.query.all():
        ws.append([c.id, c.custom_id, c.name, c.phone, c.ledger])

    ws = wb.create_sheet("Products")
    ws.append(["Code", "Name", "Main Category", "Sub Category", "Capital Price", "Cash Price"])
    for p in Product.query.all():
        ws.append([
            p.code,
            p.name,
            p.sub_category.main_category.name if p.sub_category and p.sub_category.main_category else "",
            p.sub_category.name if p.sub_category else "",
            p.capital_price,
            p.base_cash_price
        ])

    ws = wb.create_sheet("Cash Debts")
    ws.append(["Customer", "Product", "Price", "Date"])
    for d in CashDebt.query.all():
        ws.append([
            d.customer.name if d.customer else "",
            d.name,
            d.price,
            d.date_added.strftime("%Y-%m-%d") if d.date_added else ""
        ])

    ws = wb.create_sheet("Installments")
    ws.append(["Customer", "Product", "Total Price", "Initial", "Monthly", "Paid Off"])
    for i in InstallmentProduct.query.all():
        ws.append([
            i.customer.name if i.customer else "",
            i.name,
            i.total_price,
            i.initial_payment,
            i.monthly_installment,
            "Yes" if i.paid_off else "No"
        ])

    ws = wb.create_sheet("Sales")
    ws.append(["Type", "Product", "Customer", "Sell Price", "Cost Price", "Profit", "Date"])
    for s in Sale.query.all():
        ws.append([
            s.sale_type,
            s.product.name if s.product else "",
            s.customer.name if s.customer else "-",
            s.sell_price,
            s.cost_price,
            (float(s.sell_price or 0) - float(s.cost_price or 0)),
            s.date_created.strftime("%Y-%m-%d") if s.date_created else ""
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="bakkour_system_export.xlsx",
        as_attachment=True
    )
